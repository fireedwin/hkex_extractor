# -*- coding: utf-8 -*-
"""
excel_out.py — 功能C:結構化輸出(整個工具的交付成果)

設計原則(這部分最值得在面試講):

1. **每一列都帶頁碼。** 估值報告的數字要能被覆核。
   一個不能追溯來源的數字,對估值師來說等於沒有價值 —— 甚至是風險。

2. **比率用 Excel 公式,不是 Python 算完寫死。**
   分析師拿到檔案後改一個數字,整張表會自己重算。
   寫死的數字會讓他們不敢信任、也不能修改。

3. **輸出成他們本來就在用的格式。**
   不是聊天視窗裡的一段文字,是可以直接貼進工作底稿的 Excel。

4. **誠實標示不確定的地方。**
   信心度欄位、需要 OCR 的頁面清單 —— 讓人知道哪裡要人手覆核,
   比假裝 100% 準確有用得多。
"""

from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
TITLE = Font(name=FONT, bold=True, size=14, color="1F4E5F")
NOTE = Font(name=FONT, size=9, italic=True, color="808080")
BLUE = Font(name=FONT, size=10, color="0000FF")     # 硬輸入值
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)

MONEY = '#,##0;(#,##0);-'
PCT = '0.0%'


def _header(ws, row, cols, widths=None):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _style_body(ws, first_row, last_row, ncols):
    for r in range(first_row, last_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=(c > 3))


def _inject_cached_values(path: str, sheet_values: dict) -> int:
    """
    把公式的計算結果寫進 xlsx,讓不重算的軟體也能顯示數字。

    為什麼需要這個:
        openpyxl 只寫公式、不寫計算結果。Excel 通常會自動重算,
        但 LibreOffice 的「Recalculation on File Load」預設是
        「永不重算」,而且這個設定會蓋掉檔案裡的 fullCalcOnLoad 旗標。
        結果整欄顯示 #NAME? / #N/A —— 看起來像公式壞掉,
        其實只是從來沒被算過。

        把算好的值一起寫進去,兩邊都能正常顯示,
        而公式仍然保留 —— 分析師改數字時照樣會自動重算。

    sheet_values: {分頁名稱: {儲存格: 值}},值可為數字或字串。
    回傳實際寫入的儲存格數。
    """
    import zipfile
    import shutil
    import re as _re
    import html
    from xml.sax.saxutils import escape

    try:
        with zipfile.ZipFile(path) as zf:
            names = zf.namelist()
            contents = {n: zf.read(n) for n in names}
    except Exception:
        return 0

    wb_xml = contents.get("xl/workbook.xml", b"").decode("utf-8")
    rels = contents.get("xl/_rels/workbook.xml.rels", b"").decode("utf-8")

    # r:id → 檔案路徑(Target 可能帶開頭斜線)
    rid_to_target = {}
    for m in _re.finditer(r'<Relationship\b[^>]*>', rels):
        tag = m.group(0)
        rid = _re.search(r'Id="([^"]+)"', tag)
        tgt = _re.search(r'Target="([^"]+)"', tag)
        if rid and tgt:
            t = tgt.group(1).lstrip("/")
            if not t.startswith("xl/"):
                t = "xl/" + t
            rid_to_target[rid.group(1)] = t

    # 分頁名稱 → 檔案。名稱在 XML 裡是編碼過的(例如 &#26412;),要先還原
    title_to_file = {}
    for m in _re.finditer(r'<sheet\b[^>]*>', wb_xml):
        tag = m.group(0)
        name = _re.search(r'name="([^"]*)"', tag)
        rid = _re.search(r'r:id="([^"]+)"', tag)
        if name and rid:
            title_to_file[html.unescape(name.group(1))] = rid_to_target.get(rid.group(1))

    written = 0
    for title, cells in sheet_values.items():
        fn = title_to_file.get(title)
        if not fn or fn not in contents:
            continue
        xml = contents[fn].decode("utf-8")

        for ref, val in cells.items():
            if val is None:
                continue
            is_num = isinstance(val, (int, float)) and not isinstance(val, bool)

            # 公式儲存格可能有空的 <v/>、<v></v>,也可能完全沒有 <v>
            pat = _re.compile(
                r'(<c\b[^>]*\br="%s"[^>]*>)(<f>.*?</f>)(<v\s*/>|<v></v>)?(</c>)'
                % _re.escape(ref), _re.S)

            def _repl(m, val=val, is_num=is_num):
                tag, f_tag, close = m.group(1), m.group(2), m.group(4)
                if not is_num:
                    # 文字結果要標 t="str",否則會被當成錯誤代碼
                    tag = _re.sub(r'\st="[^"]*"', "", tag)
                    tag = tag[:-1] + ' t="str">'
                body = repr(val) if isinstance(val, float) else str(val)
                if is_num:
                    body = f"{val}"
                else:
                    body = escape(str(val))
                return f"{tag}{f_tag}<v>{body}</v>{close}"

            xml, n = pat.subn(_repl, xml, count=1)
            written += n

        contents[fn] = xml.encode("utf-8")

    if not written:
        return 0

    tmp = path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zf:
        for n in names:
            zf.writestr(n, contents[n])
    shutil.move(tmp, path)
    return written


# ==========================================================================
def build_workbook(out_path: str,
                   source_file: str,
                   pdf_stats: dict,
                   param_hits: list,
                   fin_result,
                   topic_hits: list,
                   ai_findings: list,
                   ocr_pages: List[int]):

    wb = Workbook()
    # 一邊寫公式、一邊用 Python 算出同樣的結果,最後注入成快取值,
    # 讓不自動重算的軟體(LibreOffice 預設)也能直接看到數字。
    cached = {}

    # ---------------- Sheet 1: 摘要 -------------------------------------
    ws = wb.active
    ws.title = "Summary 摘要"
    ws["A1"] = "HKEX 文件資料萃取報告"
    ws["A1"].font = TITLE
    ws["A2"] = "Automated Extraction Report — Business Valuation Support"
    ws["A2"].font = NOTE

    rows = [
        ("來源文件 Source file", source_file),
        ("產生時間 Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("總頁數 Total pages", pdf_stats.get("total_pages", 0)),
        ("已擷取字元 Characters extracted", pdf_stats.get("total_chars", 0)),
        ("含表格頁數 Pages with tables", pdf_stats.get("pages_with_tables", 0)),
        ("需人手/OCR 頁數 Pages needing OCR", pdf_stats.get("pages_needing_ocr", 0)),
        ("", ""),
        ("估值參數擷取筆數 Valuation parameters", len(param_hits)),
        ("財務科目擷取筆數 Financial line items", len(fin_result.items)),
        ("主題段落擷取筆數 Topic extracts", len(topic_hits)),
        ("AI 複核筆數 AI-verified findings", len(ai_findings)),
    ]
    r = 4
    for k, v in rows:
        ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=bool(k), size=10)
        ws.cell(row=r, column=2, value=v).font = BODY
        r += 1

    ws.cell(row=r + 1, column=1, value="資料完整性檢查 Integrity checks").font = Font(
        name=FONT, bold=True, size=11)
    r += 2
    ws.cell(row=r, column=1,
            value="用會計恆等式交叉驗證擷取結果 —— 數字之間本來就該對得上,"
                  "對不上代表某欄可能抓錯,請翻回來源頁核對。").font = NOTE
    r += 1
    try:
        from financials import integrity_checks
        checks = integrity_checks(fin_result)
    except Exception:
        checks = []
    if checks:
        for name, passed, detail in checks:
            c = ws.cell(row=r, column=1, value=("✓ " if passed else "✗ ") + name)
            c.font = Font(name=FONT, size=10,
                          color="006100" if passed else "9C0006", bold=not passed)
            ws.cell(row=r, column=2, value=detail).font = BODY
            r += 1
    else:
        ws.cell(row=r, column=1,
                value="擷取到的科目不足,無法進行交叉驗證").font = BODY
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="方法論 Methodology").font = Font(name=FONT, bold=True, size=11)
    method = [
        "1. 逐頁擷取文字,全程保留 PDF 頁碼與年報印刷頁碼(可追溯性)。",
        "2. 規則層(正則/關鍵字)先把全份文件縮減為少數相關頁面 —— 解決「文件太大」。",
        "3. 數字由正則精確擷取並做合理性檢查,不由語言模型生成 —— 解決「準確度」。",
        "4. AI 僅在縮減後的頁面上做語意判斷,且輸出頁碼須通過驗證才保留。",
        "5. 所有結果附來源頁碼與原文行,供估值師覆核。信心度為 Low 者建議人手確認。",
        "6. 支援中英對照年報、破折號(nil)、Profit/(loss) 等港股常見排版。",
    ]
    for i, m in enumerate(method):
        ws.cell(row=r + 1 + i, column=1, value=m).font = BODY
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 60

    # ---------------- Sheet 2: 估值參數 ---------------------------------
    ws = wb.create_sheet("Valuation Params 估值參數")
    ws["A1"] = "估值參數 Benchmark — 可直接併入行內參數資料庫"
    ws["A1"].font = TITLE
    ws["A2"] = "數字由正則式從原文精確擷取,並經合理範圍檢查;非由 AI 生成。"
    ws["A2"].font = NOTE

    cols = ["參數 Parameter", "中文", "下限 Low (%)", "上限 High (%)",
            "原文數值", "信心度", "來源頁 Source Page", "上下文 Context"]
    _header(ws, 4, cols, [24, 18, 12, 12, 14, 10, 22, 90])

    r = 5
    for h in param_hits:
        ws.cell(row=r, column=1, value=h.parameter)
        ws.cell(row=r, column=2, value=h.parameter_zh)
        ws.cell(row=r, column=3, value=h.value_low).number_format = '0.00'
        ws.cell(row=r, column=4, value=h.value_high).number_format = '0.00'
        ws.cell(row=r, column=5, value=h.raw_text)
        c = ws.cell(row=r, column=6, value=h.confidence)
        if h.confidence == "Low":
            c.fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=r, column=7, value=h.page_cite)
        ws.cell(row=r, column=8, value=h.context)
        r += 1
    if r > 5:
        _style_body(ws, 5, r - 1, len(cols))

    # ---------------- Sheet 3: 財務數據 ---------------------------------
    ws = wb.create_sheet("Financials 財務數據")
    ws["A1"] = "三大報表擷取結果"
    ws["A1"].font = TITLE
    year_note = "; ".join(f"{k}: {v}" for k, v in fin_result.year_labels.items() if v)
    ws["A2"] = f"偵測到的年度欄位 — {year_note or '未偵測'}"
    ws["A2"].font = NOTE

    cols = ["報表 Statement", "科目 Item", "本年 Current", "上年 Prior",
            "變動 YoY %", "來源頁", "原文行 Source line"]
    _header(ws, 4, cols, [20, 28, 16, 16, 12, 20, 70])

    r = 5
    first_fin = r
    for it in fin_result.items:
        ws.cell(row=r, column=1, value=it.statement)
        ws.cell(row=r, column=2, value=it.item)
        c1 = ws.cell(row=r, column=3, value=it.current_year); c1.number_format = MONEY; c1.font = BLUE
        c2 = ws.cell(row=r, column=4, value=it.prior_year);   c2.number_format = MONEY; c2.font = BLUE
        # YoY 用公式,不是 Python 算完寫死 —— 分析師改數字時會自動重算
        # 刻意不用 IFERROR:實測有些 LibreOffice 版本不認得這個函式,
        # 會整欄顯示 #NAME?。ISERROR 是最古老的寫法,所有試算表都支援。
        _yoy = f'(C{r}-D{r})/ABS(D{r})'
        f = ws.cell(row=r, column=5,
                    value=f'=IF(ISERROR({_yoy}),"",{_yoy})')
        f.number_format = PCT
        cached.setdefault(ws.title, {})[f"E{r}"] = (
            "" if not it.prior_year
            else (it.current_year - it.prior_year) / abs(it.prior_year))
        ws.cell(row=r, column=6, value=it.page_cite)
        ws.cell(row=r, column=7, value=it.source_line)
        r += 1
    last_fin = r - 1
    if last_fin >= first_fin:
        _style_body(ws, first_fin, last_fin, len(cols))
        ws.cell(row=last_fin + 2, column=1,
                value="藍色 = 從 PDF 原文擷取的數值(硬輸入);黑色 = 公式計算。"
                      "單位依原報表所示(通常為千元或百萬元),請對照來源頁確認。").font = NOTE

    # ---------------- Sheet 4: 財務比率(全公式) ------------------------
    ws = wb.create_sheet("Ratios 財務比率")
    ws["A1"] = "衍生比率 — 全部以公式連結至 Financials 分頁"
    ws["A1"].font = TITLE
    ws["A2"] = "改動 Financials 分頁的數字,本頁會自動重算。"
    ws["A2"].font = NOTE

    _header(ws, 4, ["比率 Ratio", "本年 Current", "上年 Prior", "說明"],
            [30, 16, 16, 48])

    FIN = "'Financials 財務數據'"

    def lookup(item, col):
        """
        用 INDEX/MATCH 從 Financials 分頁取值。

        避開 XLOOKUP(較新)也避開 IFERROR —— 實測有些 LibreOffice
        版本不認得 IFERROR,會讓整欄變成 #NAME?。
        錯誤處理改用外層的 IF(ISERROR(...)),相容性最高。
        """
        lo, hi = first_fin, max(last_fin, first_fin)
        return (f'INDEX({FIN}!${col}${lo}:${col}${hi},'
                f'MATCH("{item}",{FIN}!$B${lo}:$B${hi},0))')

    ratios = [
        ("Gross Margin 毛利率", "Gross Profit", "Revenue", PCT, "毛利 ÷ 收入"),
        ("Net Margin 淨利率", "Profit for the Year", "Revenue", PCT, "年內溢利 ÷ 收入"),
        ("Current Ratio 流動比率", "Total Current Assets", "Total Current Liabilities", '0.00', "流動資產 ÷ 流動負債"),
        ("Debt / Equity 負債權益比", "Total Liabilities", "Total Equity", '0.00', "總負債 ÷ 總權益"),
        ("Goodwill / Total Assets", "Goodwill", "Total Assets", PCT, "商譽佔總資產比重 — 減值風險指標"),
        ("Intangibles / Total Assets", "Intangible Assets", "Total Assets", PCT, "無形資產佔比"),
        ("ROE 股本回報率", "Profit for the Year", "Total Equity", PCT, "年內溢利 ÷ 總權益"),
    ]
    # 用 Python 重算一次比率,供快取值注入
    by_item = {}
    for it in fin_result.items:
        by_item[it.item] = it

    def _val(item, which):
        it = by_item.get(item)
        if it is None:
            return None
        return it.current_year if which == "C" else it.prior_year

    r = 5
    for label, num, den, fmt, desc in ratios:
        ws.cell(row=r, column=1, value=label).font = BODY
        for ci, col in ((2, "C"), (3, "D")):
            expr = f'{lookup(num, col)}/{lookup(den, col)}'
            cell = ws.cell(row=r, column=ci,
                           value=f'=IF(ISERROR({expr}),"n/a",{expr})')
            cell.number_format = fmt
            cell.font = BODY
            n_v, d_v = _val(num, col), _val(den, col)
            cached.setdefault(ws.title, {})[f"{get_column_letter(ci)}{r}"] = (
                n_v / d_v if (n_v is not None and d_v) else "n/a")
        ws.cell(row=r, column=4, value=desc).font = NOTE
        r += 1
    ws.cell(row=r + 1, column=1,
            value='顯示 "n/a" 表示該科目未在文件中成功擷取,需人手補入。').font = NOTE

    # ---------------- Sheet 5: 趨勢圖 -----------------------------------
    if fin_result.items:
        ws = wb.create_sheet("Chart 趨勢圖")
        ws["A1"] = "本年 vs 上年 — 主要科目比較"
        ws["A1"].font = TITLE
        ws["A2"] = ("港交所主板及 GEM 自 2024 年起均無強制季度報告,"
                    "故以年度 / 中期數據作趨勢比較較符合披露實況。")
        ws["A2"].font = NOTE

        key_items = ["Revenue", "Gross Profit", "Profit Before Tax",
                     "Profit for the Year", "Total Assets", "Total Equity"]
        by_name = {i.item: i for i in fin_result.items}
        ws.cell(row=4, column=1, value="Item").font = HDR_FONT
        ws.cell(row=4, column=2, value="Prior").font = HDR_FONT
        ws.cell(row=4, column=3, value="Current").font = HDR_FONT
        for c in range(1, 4):
            ws.cell(row=4, column=c).fill = HDR_FILL

        rr = 5
        for name in key_items:
            it = by_name.get(name)
            if not it or it.current_year is None:
                continue
            ws.cell(row=rr, column=1, value=name).font = BODY
            ws.cell(row=rr, column=2, value=it.prior_year).number_format = MONEY
            ws.cell(row=rr, column=3, value=it.current_year).number_format = MONEY
            rr += 1

        if rr > 5:
            ch = BarChart()
            ch.type, ch.style = "col", 10
            ch.title = "Key Financials — Prior vs Current"
            ch.y_axis.title = "Amount (as reported)"
            data = Reference(ws, min_col=2, max_col=3, min_row=4, max_row=rr - 1)
            cats = Reference(ws, min_col=1, min_row=5, max_row=rr - 1)
            ch.add_data(data, titles_from_data=True)
            ch.set_categories(cats)
            ch.width, ch.height = 22, 11
            ws.add_chart(ch, "E4")

    # ---------------- Sheet 6: 主題段落 ---------------------------------
    ws = wb.create_sheet("Extracts 主題段落")
    ws["A1"] = "主題式段落擷取 — 供估值師快速定位原文"
    ws["A1"].font = TITLE
    _header(ws, 3, ["主題 Topic", "中文", "優先", "命中詞", "來源頁", "段落 Snippet"],
            [28, 20, 10, 20, 20, 100])
    r = 4
    for h in sorted(topic_hits, key=lambda x: ({"high": 0, "medium": 1, "low": 2}[x.priority],
                                               x.topic, x.page_index)):
        ws.cell(row=r, column=1, value=h.topic)
        ws.cell(row=r, column=2, value=h.topic_zh)
        c = ws.cell(row=r, column=3, value=h.priority)
        if h.priority == "high":
            c.fill = PatternFill("solid", fgColor="E2EFDA")
        ws.cell(row=r, column=4, value=h.matched_term)
        ws.cell(row=r, column=5, value=h.page_cite)
        ws.cell(row=r, column=6, value=h.snippet)
        r += 1
    if r > 4:
        _style_body(ws, 4, r - 1, 6)

    # ---------------- Sheet 7: AI 複核 ----------------------------------
    if ai_findings:
        ws = wb.create_sheet("AI Review AI複核")
        ws["A1"] = "AI 語意層擷取結果(頁碼已驗證)"
        ws["A1"].font = TITLE
        ws["A2"] = "僅在規則層篩出的頁面上執行;模型回報的頁碼須存在於送入範圍才保留。"
        ws["A2"].font = NOTE
        _header(ws, 4, ["類別", "項目", "數值", "單位", "用途 Used for", "頁碼", "原文引述"],
                [22, 30, 18, 12, 34, 10, 46])
        r = 5
        for f in ai_findings:
            for ci, k in enumerate(["category", "item", "value", "unit",
                                    "used_for", "page", "verbatim"], 1):
                ws.cell(row=r, column=ci, value=f.get(k, ""))
            r += 1
        _style_body(ws, 5, r - 1, 7)

    # ---------------- Sheet 8: 待人手處理 -------------------------------
    ws = wb.create_sheet("Review Queue 待覆核")
    ws["A1"] = "需要人手處理的項目"
    ws["A1"].font = TITLE
    ws["A2"] = "工具沒抓到的東西會列在這裡,不會靜靜消失 —— 請人手翻閱原文補入。"
    ws["A2"].font = NOTE

    r = 4
    # ── 未擷取到的財務科目 ──────────────────────────
    missing = getattr(fin_result, "missing", {}) or {}
    ws.cell(row=r, column=1, value="一、未擷取到的財務科目").font = Font(
        name=FONT, bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1,
            value="可能原因:該公司未單獨揭露此科目 / 用了工具不認得的措辭 / "
                  "位於跨頁或附註中。措辭問題可在 config.py 補上別名。").font = NOTE
    r += 2

    if missing:
        _header(ws, r, ["報表 Statement", "未擷取科目 Missing item"], [24, 40])
        r += 1
        first = r
        for stmt, items in missing.items():
            for it in items:
                ws.cell(row=r, column=1, value=stmt)
                ws.cell(row=r, column=2, value=it)
                r += 1
        _style_body(ws, first, r - 1, 2)
    else:
        ws.cell(row=r, column=1, value="無 — 所有設定的財務科目均已擷取").font = BODY
        r += 1

    # ── 需要 OCR 的頁面 ─────────────────────────────
    r += 2
    ws.cell(row=r, column=1, value="二、文字量過低的頁面(疑似掃描頁)").font = Font(
        name=FONT, bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1,
            value="這些頁面幾乎抓不到文字,通常是掃描件或純圖表,"
                  "建議人手翻閱或另行 OCR。").font = NOTE
    r += 2
    _header(ws, r, ["PDF 頁碼", "狀態"], [16, 40])
    r += 1
    if ocr_pages:
        first = r
        for p in ocr_pages:
            ws.cell(row=r, column=1, value=p)
            ws.cell(row=r, column=2, value="文字量過低 — 可能為掃描頁/圖表頁")
            r += 1
        _style_body(ws, first, r - 1, 2)
    else:
        ws.cell(row=r, column=1, value="無 — 全份文件均成功擷取文字").font = BODY

    # ── 強制開啟時重算 ─────────────────────────────────
    # openpyxl 只寫公式、不寫計算結果,留下空的 <v/> 快取值。
    # LibreOffice(以及部分 Excel 設定)開啟 xlsx 時預設不重算,
    # 讀到空快取就把整欄顯示成 #NAME? / #N/A —— 看起來像公式壞掉,
    # 其實公式完全正確,只是從來沒被算過。
    # 這個旗標會告訴試算表軟體「開檔時請全部重算一次」。
    wb.calculation.fullCalcOnLoad = True

    wb.save(out_path)

    # 注入快取值 —— 失敗也不影響檔案本身(公式仍在,只是要靠軟體重算)
    try:
        _inject_cached_values(out_path, cached)
    except Exception as e:
        print(f"[excel] 快取值注入失敗(不影響公式): {e}")

    return out_path
